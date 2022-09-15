"""
Author:     Christopher Lien
Date:       20220915
Version:    1.1
"""

import matplotlib.pyplot as plt
import os
import sys
import numpy as np
import openpyxl as py
from PIL import Image
from math import sqrt
from tkinter import filedialog


class Main():
    def __init__(self, ):
        self.xlsm_dir = self.get_xlsm_location()
        image = "piper_background.png"
        image_dir = self.resource_path(image)
        self.image = Image.open(image_dir)
        self.header = ['Bore_ID','Color_ID','HCO3','CO3','SO4','Cl','Na','Ca',
                       'Mg','K', 'Symbol_ID', 'Size']
        self.ions = {'HCO3':61,'CO3':30,'Cl':35,'SO4':48,'Na':23,'Ca':20,
                     'Mg':12,'K':39}

    def get_xlsm_location(self, ):
        xlsm_dir = filedialog.askopenfilename(initialdir="/",
                                              title="Select file",
                                              filetypes=(("piper data", "*.xlsm"),
                                                         ("all files", "*.*"))
                                              )

        return xlsm_dir

    def resource_path(self, image):
        """Get absolute path to resource, works for dev and for PyInstaller
        """
        path = os.path.dirname(os.path.abspath(__file__))
        base_path = getattr(sys, '_MEIPASS', path)

        return os.path.join(base_path, image)

    def build_meq_dict(self, ):
        """Read xlsm file and parse data to list[dict{}]
        """
        workbook = py.load_workbook(self.xlsm_dir, data_only=True)
        data = workbook.active
        header = []
        list_meq = []
        for col in range(1, data.max_column+1):
            header.append(data.cell(row=1,column=col).value)
        for row in range (2, data.max_row):
            dict_meq = {}
            for col in range(1, data.max_column+1):
                dict_meq[header[col-1]]=data.cell(row=row,column=col).value
            list_meq.append(dict_meq)

        """Create list[dict{}] of header data and convert from mg/L to meq
        """
        dict_list = []
        for i in list_meq:
            dict_meq = {}
            for key, value in i.items():
                for j in self.header:
                    if key==j:
                        try:
                            meq_val = self.ions[key]
                            value = value/meq_val
                            dict_meq[key] = value
                        except:
                            dict_meq[key] = value
            dict_list.append(dict_meq)
        self.normalise(dict_list)

    def normalise(self, list_meq):
        """Generate new list[dict{}] of parsed xlsm data to normalised values
        for generating x,y coordinates for plotting.

        Args:
            list_meq (list): list[dict{}] of xlsm data
        """
        list_norm = []
        for i in list_meq:
            dict_norm = {}
            dict_norm['Bore_ID'] = i['Bore_ID']
            dict_norm['Color_ID'] = i['Color_ID']
            dict_norm['Symbol_ID'] = i['Symbol_ID']
            dict_norm['Size'] = i['Size']
            dict_norm['SO4'] = i['SO4'] / (
                i['SO4'] + i['HCO3'] + i['CO3'] + i['Cl']
                ) * 100
            dict_norm['HCO3_CO3'] = (i['HCO3'] + i['CO3']) / (
                i['SO4'] + i['HCO3'] + i['CO3'] + i['Cl']
                ) * 100
            dict_norm['Cl'] = i['Cl'] / (
                i['SO4'] + i['HCO3'] +i['CO3'] +i['Cl']
                ) * 100
            dict_norm['Mg'] = i['Mg'] / (
                i['Mg'] + i['Ca'] + i['K'] + i['Na']
                ) * 100
            dict_norm['Na_K'] = (i['K'] + i['Na']) / (
                i['Mg'] + i['Ca'] + i['K'] + i['Na']
                ) * 100
            dict_norm['Ca'] = i['Ca'] / (
                i['Mg'] + i['Ca']+i['K']+i['Na']
                ) * 100
            list_norm.append(dict_norm)
        list_norm = self.hide_legend(list_norm)
        self.display(list_norm)

    def hide_legend(self, list_norm):
        dup_list = []
        for i in list_norm:
            if i['Bore_ID'] in dup_list:
                i['Bore_ID'] = '_'+i['Bore_ID']
            else:
                dup_list.append(i['Bore_ID'])
        return list_norm

    def display(self, list_norm):
        """Display method for parsing data through matplotlib, save file once
        complete.

        Args:
            list_norm (list): list[dict{}] of normalised xlsm data for plotting.
        """
        plt.figure(figsize=(20,15))
        plt.imshow(np.flipud(self.image))
        for i in list_norm:
            self.coordinate(i['Ca'], i['Mg'], i['Cl'], i['SO4'], i['Bore_ID'],
                            i['Color_ID'], i['Symbol_ID'], i['Size'])
        plt.ylim(0,830)
        plt.xlim(0,900)
        plt.axis('off')
        plt.legend(loc='upper right',prop={'size':10}, frameon=False,
                   scatterpoints=1)
        file_loc = self.xlsm_dir.replace(".xlsm", "")
        sav_loc_png = file_loc+"_piper_plot.png"
        plt.savefig(sav_loc_png)
        print("Output Complete to %s. Exiting." % (sav_loc_png))
        exit

    def coordinate(self, ca, mg, cl, so4, label, color, marker, size):
        """Convert data to x,y coordinates for plotting through matplotlib.

        Args:
            Ca (float): Ca as meq
            Mg (float): Mg as meq
            Cl (float): Cl as meq
            SO4 (float): SO4 as meq
            Label (string): Groundwater bore identification name
            Color (string): Groundwater bore colour for plotting
            Marker (String): Groundwater bore marker for plotting
            Size (int): Groundwater bore size for plotting

        Returns:
            list: list for plotting piper plot through matplotlib
        """
        list_coord = []
        edge_col = '#4b4b4b'
        zorder = 1
        xcation = 40 + 360 - (ca + mg / 2) * 3.6
        ycation = 40 + (sqrt(3) * mg / 2) * 3.6
        xanion = 40 + 360 + 100 + (cl + so4 / 2) * 3.6
        yanion = 40 + (so4 * sqrt(3) / 2) * 3.6
        xdiam = 0.5 * (xcation + xanion + (yanion - ycation) / sqrt(3))
        ydiam = 0.5 * (yanion + ycation + sqrt(3) * (xanion - xcation))
        list_coord.append(plt.scatter(xcation,ycation,zorder=zorder,c=color,
                                      s=size, edgecolors=edge_col,label=label,
                                      marker=marker))
        list_coord.append(plt.scatter(xanion,yanion,zorder=zorder,c=color,
                                      s=size, edgecolors=edge_col,
                                      marker=marker))
        list_coord.append(plt.scatter(xdiam,ydiam,zorder=zorder,c=color,
                                      s=size, edgecolors=edge_col,
                                      marker=marker))
        return list_coord

if __name__=="__main__":
    Main().build_meq_dict()
